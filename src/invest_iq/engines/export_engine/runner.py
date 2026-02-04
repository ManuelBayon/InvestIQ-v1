from pathlib import Path

from openpyxl import load_workbook

from invest_iq.engines.backtest_engine.common.types import ExecutionLogEntry
from invest_iq.engines.export_engine.factory import ExportServiceFactory
from invest_iq.engines.export_engine.registries.config import ExportKey, ExportOptions

from invest_iq.engines.utilities.logger.factory import LoggerFactory


class BacktestExportRunner:

    def __init__(
            self,
            logger_factory: LoggerFactory,
            export_key: ExportKey,
            export_options: ExportOptions,
    ):
        self._logger_factory = logger_factory
        self._export_key = export_key
        self._export_options = export_options


        self._export_service_factory = ExportServiceFactory(
            logger_factory=self._logger_factory,
            options=self._export_options,
        )

    def export(
        self,
        execution_log: list[ExecutionLogEntry],
        metrics: dict[str, float] | None = None,
    ) -> None:

        export_service = self._export_service_factory.create_backtest_batch_export_service(
            key=self._export_key,
            logger_factory=self._logger_factory,
        )

        export_path: Path = export_service.export(raw_data=execution_log)

        self._append_metrics_excel(
            path=export_path,
            metrics=metrics or {
                "Realized Pnl": metrics["Realized Pnl"],
            },
        )
    @staticmethod
    def _append_metrics_excel(path: Path, metrics: dict[str, float]) -> None:
        wb = load_workbook(path)

        # 1) Create or replace "Metrics" sheet
        sheet_name = "Metrics"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # clear existing content (simple wipe)
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        # 2) Write key-value table
        ws["B1"] = "Metric"
        ws["C1"] = "Value"

        row = 2
        for k, v in metrics.items():
            ws[f"B{row}"] = str(k)
            ws[f"C{row}"] = float(v)
            row += 1

        wb.save(path)
        wb.close()